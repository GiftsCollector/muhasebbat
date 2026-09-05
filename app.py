from collections import defaultdict, deque
from datetime import date, datetime
from io import BytesIO
import json
import os
import secrets
import zipfile

from flask import Flask, render_template, request, url_for, flash, redirect, session, g, send_file, abort
from flask_wtf.csrf import CSRFProtect
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.utils import secure_filename

from models import (
    db, User, Project, BOQItem, ChartOfAccount, ProgressPayment, ProgressPaymentItem,
    CostEntry, Subcontractor, Supplier, PurchaseOrder, InventoryTransaction, LaborEntry,
    Equipment, JournalEntry, CustodySettlement, DriverCompensationEntry, SubcontractorPayment,
    Estimation, EstimationItem, ClientReceipt, SupplierPayment, AccountingPeriod,
    DocumentAttachment, ActivityLog, Employee, EmployeeAttendance, PayrollSlip,
    EmployeeSalaryPayment, ROLE_ADMIN, ROLE_ACCOUNTANT, ROLE_DATA_ENTRY, ROLE_PROJECT_MANAGER,
    ROLE_LABELS,
)
from services.accounting import (
    PeriodClosedError, as_float, as_int, assert_period_open, build_account_balances,
    build_custody_balances, build_custody_owner_rows, build_project_cost_breakdown,
    build_subcontractor_statement, calculate_purchase_order_total, classify_balance_sheet_section,
    client_progress_query, document_journal_markers, ensure_core_accounting_accounts,
    find_closed_period, format_grouped_number, generate_estimation_code,
    generate_progress_payment_number, get_account_by_code, get_age_bucket,
    get_custody_entity_accounts, get_custody_expense_accounts, get_custody_operation_type,
    get_custody_owner_type, get_custody_settlement_amount, get_journal_entries_in_range,
    get_main_treasury_rollup_balance, get_material_names, get_or_create_client_account,
    get_or_create_custody_owner_account, get_or_create_subcontractor_account,
    get_or_create_supplier_account, get_subcontractor_outstanding_advances,
    get_treasury_balance, get_warehouse_names, is_auto_journal_entry,
    is_operational_treasury_account, is_purchase_order_inventory, is_treasury_account,
    list_client_names, normalize_entity_kinds, normalize_journal_status,
    parse_custody_settlement_lines, parse_estimation_items, parse_progress_payment_items,
    recalculate_estimation, recalculate_progress_payment, replace_auto_journals,
    resolve_client_name, run_schema_migrations, split_treasury_accounts, subcontractor_progress_query,
    sync_client_receipt_journal, sync_custody_settlement_journal,
    sync_driver_compensation_journals, sync_equipment_journals, sync_estimation_journal,
    sync_inventory_transaction_journal, sync_journal_related_accounts, sync_labor_journal,
    sync_progress_payment_journals, sync_purchase_order_journal,
    sync_purchase_order_to_inventory, sync_subcontractor_payment_journal,
    sync_supplier_payment_journal, void_document_journals,
    JOURNAL_OPTIONS, ENTITY_KIND_OPTIONS, UNIT_OPTIONS, EXPENSE_CLASS_OPTIONS,
    CUSTODY_EXPENSE_NATURES, CUSTODY_OPERATION_TYPES, CUSTODY_OWNER_TYPES,
    CUSTODY_DISBURSE_OPERATIONS,
)
from services.authz import (
    ROLE_OPTIONS, current_perms, endpoint_allowed, normalize_legacy_role, require_perm, user_can,
)
from services.audit import register_audit_hooks

app = Flask(__name__)

_WEAK_SECRET_KEYS = {
    "",
    "dev-secret-key-2024",
    "dev-only-change-me",
    "your-secret-key-here-change-in-production",
}


def _resolve_secret_key():
    env_key = (os.getenv("SECRET_KEY") or "").strip()
    if env_key not in _WEAK_SECRET_KEYS:
        return env_key

    instance_dir = os.path.join(os.path.dirname(__file__), "instance")
    os.makedirs(instance_dir, exist_ok=True)
    secret_path = os.path.join(instance_dir, ".secret_key")
    if os.path.isfile(secret_path):
        try:
            with open(secret_path, encoding="utf-8") as handle:
                stored = handle.read().strip()
            if stored not in _WEAK_SECRET_KEYS:
                return stored
        except OSError:
            pass

    generated = secrets.token_hex(32)
    try:
        with open(secret_path, "w", encoding="utf-8") as handle:
            handle.write(generated)
    except OSError:
        pass
    print(
        "WARNING: SECRET_KEY is missing. Generated a temporary key. "
        "Set SECRET_KEY in Render Environment so logins survive deploys."
    )
    return generated


DATABASE_URL = (os.getenv("DATABASE_URL") or "").strip()
IS_PRODUCTION = bool(DATABASE_URL or os.getenv("RENDER") or os.getenv("FLASK_ENV") == "production")
app.config["SECRET_KEY"] = _resolve_secret_key()
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024
app.config["WTF_CSRF_TIME_LIMIT"] = None

if DATABASE_URL:
    if DATABASE_URL.startswith("postgres://"):
        DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"pool_pre_ping": True, "pool_recycle": 280}
else:
    db_folder = os.path.join(os.path.dirname(__file__), "instance")
    os.makedirs(db_folder, exist_ok=True)
    db_path = os.path.join(db_folder, "data.db")
    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{db_path}"
    if IS_PRODUCTION:
        print(
            "WARNING: DATABASE_URL is missing. Using temporary SQLite. "
            "Create a Render PostgreSQL database and set DATABASE_URL or data will be lost on deploy."
        )

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "instance", "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

db.init_app(app)
csrf = CSRFProtect(app)
app.jinja_env.filters["groupnum"] = format_grouped_number
register_audit_hooks()

with app.app_context():
    db.create_all()
    run_schema_migrations()
    normalize_entity_kinds()

PUBLIC_ENDPOINTS = {"login", "setup_admin", "static"}
PASSWORD_CHANGE_ENDPOINTS = {"change_password", "logout"}


@app.before_request
def load_user_and_protect_routes():
    g.current_user = None
    user_id = session.get("user_id")
    if user_id:
        g.current_user = User.query.get(user_id)
        if g.current_user is None or not g.current_user.is_active:
            session.clear()

    endpoint = request.endpoint or ""
    if endpoint in PUBLIC_ENDPOINTS or endpoint.startswith("static"):
        return None

    if g.current_user is None:
        return redirect(url_for("login", next=request.path))

    if session.get("must_change_password") and endpoint not in PASSWORD_CHANGE_ENDPOINTS:
        flash("يجب تغيير كلمة المرور الافتراضية قبل استخدام النظام", "warning")
        return redirect(url_for("change_password"))

    if not endpoint_allowed(endpoint):
        flash("ليست لديك صلاحية للوصول إلى هذه الشاشة", "danger")
        return redirect(url_for("index"))

    return None


@app.context_processor
def inject_current_user():
    user = getattr(g, "current_user", None)
    return {
        "current_user": user,
        "is_logged_in": user is not None,
        "is_admin": bool(user and user.role == ROLE_ADMIN),
        "user_can": user_can,
        "can_endpoint": endpoint_allowed,
        "role_labels": ROLE_LABELS,
    }


@app.route("/login", methods=["GET", "POST"])
def login():
    if User.query.count() == 0:
        return redirect(url_for("setup_admin"))

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        user = User.query.filter_by(username=username).first()

        if user and user.is_active and user.check_password(password):
            session["user_id"] = user.id
            if user.check_password("123456"):
                session["must_change_password"] = True
                flash("كلمة المرور الحالية غير آمنة. غيّرها قبل متابعة العمل.", "warning")
                return redirect(url_for("change_password"))
            flash("تم تسجيل الدخول بنجاح", "success")
            next_url = request.args.get("next")
            if next_url and next_url.startswith("/"):
                return redirect(next_url)
            return redirect(url_for("index"))

        flash("بيانات الدخول غير صحيحة", "danger")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("تم تسجيل الخروج", "success")
    return redirect(url_for("login"))


@app.route("/setup-admin", methods=["GET", "POST"])
def setup_admin():
    if User.query.count() > 0:
        return redirect(url_for("login"))

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        full_name = (request.form.get("full_name") or "").strip()
        password = request.form.get("password") or ""
        confirm_password = request.form.get("confirm_password") or ""

        if not username or not full_name or not password:
            flash("يرجى استكمال جميع الحقول", "danger")
            return redirect(url_for("setup_admin"))

        if len(password) < 6:
            flash("كلمة المرور يجب أن تكون 6 أحرف على الأقل", "danger")
            return redirect(url_for("setup_admin"))

        if password == "123456":
            flash("لا تستخدم كلمة المرور الافتراضية 123456. اختر كلمة أقوى.", "danger")
            return redirect(url_for("setup_admin"))

        if password != confirm_password:
            flash("تأكيد كلمة المرور غير متطابق", "danger")
            return redirect(url_for("setup_admin"))

        admin = User(username=username, full_name=full_name, role=ROLE_ADMIN, is_active=True)
        admin.set_password(password)
        db.session.add(admin)
        db.session.commit()

        flash("تم إنشاء حساب المدير بنجاح. يمكنك تسجيل الدخول الآن", "success")
        return redirect(url_for("login"))

    return render_template("setup_admin.html")


@app.route("/users", methods=["GET", "POST"])
def users():
    if not user_can("*") and g.current_user.role != ROLE_ADMIN:
        flash("ليس لديك صلاحية للوصول لإدارة المستخدمين", "danger")
        return redirect(url_for("index"))

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        full_name = (request.form.get("full_name") or "").strip()
        password = request.form.get("password") or ""
        role = normalize_legacy_role(request.form.get("role") or ROLE_DATA_ENTRY)
        is_active = request.form.get("is_active") == "on"

        if not username or not full_name or not password:
            flash("يرجى استكمال بيانات المستخدم", "danger")
            return redirect(url_for("users"))

        if password == "123456":
            flash("لا تستخدم كلمة المرور الافتراضية 123456", "danger")
            return redirect(url_for("users"))

        if User.query.filter_by(username=username).first():
            flash("اسم المستخدم مستخدم بالفعل", "danger")
            return redirect(url_for("users"))

        user = User(username=username, full_name=full_name, role=role, is_active=is_active)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash("تم إنشاء المستخدم بنجاح", "success")
        return redirect(url_for("users"))

    all_users = User.query.order_by(User.id.desc()).all()
    return render_template("users.html", users=all_users, role_options=ROLE_OPTIONS)


@app.route("/activity-log")
def activity_log():
    if not user_can("*") and g.current_user.role != ROLE_ADMIN:
        flash("سجل الحركات متاح لمدير النظام فقط", "danger")
        return redirect(url_for("index"))

    from_date = (request.args.get("from_date") or "").strip()
    to_date = (request.args.get("to_date") or "").strip()
    user_id = (request.args.get("user_id") or "").strip()
    action = (request.args.get("action") or "").strip()
    entity_label = (request.args.get("entity_label") or "").strip()

    query = ActivityLog.query
    if from_date:
        query = query.filter(ActivityLog.created_at >= from_date + " 00:00:00")
    if to_date:
        query = query.filter(ActivityLog.created_at <= to_date + " 23:59:59")
    if user_id.isdigit():
        query = query.filter(ActivityLog.user_id == int(user_id))
    if action:
        query = query.filter(ActivityLog.action == action)
    if entity_label:
        query = query.filter(ActivityLog.entity_label == entity_label)

    rows = query.order_by(ActivityLog.id.desc()).limit(1000).all()
    users = User.query.order_by(User.full_name, User.username).all()
    actions = ["إضافة", "تعديل", "حذف", "ترحيل", "إلغاء ترحيل"]
    entity_labels = sorted({label for label in (
        "قيد يومية", "مستخلص", "دفعة مقاول باطن", "تحصيل عميل", "سداد مورد",
        "أمر شراء", "حركة مخزون", "حركة عمالة", "معدة", "تسوية عهدة",
        "محاسبة سائق", "مقايسة", "قيد تكلفة", "مشروع", "حساب", "مورد", "مقاول باطن",
    )})
    return render_template(
        "activity_log.html",
        rows=rows,
        users=users,
        actions=actions,
        entity_labels=entity_labels,
        from_date=from_date,
        to_date=to_date,
        selected_user_id=user_id,
        selected_action=action,
        selected_entity_label=entity_label,
    )


@app.route("/users/<int:user_id>/update", methods=["POST"])
def update_user(user_id):
    if g.current_user is None or g.current_user.role != ROLE_ADMIN:
        flash("ليس لديك صلاحية لتعديل المستخدمين", "danger")
        return redirect(url_for("index"))

    user = User.query.get_or_404(user_id)
    user.full_name = (request.form.get("full_name") or "").strip() or user.full_name
    user.role = normalize_legacy_role(request.form.get("role") or user.role)
    user.is_active = request.form.get("is_active") == "on"

    new_password = request.form.get("password") or ""
    if new_password:
        if len(new_password) < 6:
            flash("كلمة المرور الجديدة يجب أن تكون 6 أحرف على الأقل", "danger")
            return redirect(url_for("users"))
        if new_password == "123456":
            flash("لا يمكن تعيين كلمة المرور الافتراضية 123456", "danger")
            return redirect(url_for("users"))
        user.set_password(new_password)

    if user.id == g.current_user.id and not user.is_active:
        flash("لا يمكنك تعطيل حسابك أثناء تسجيل الدخول", "danger")
        return redirect(url_for("users"))

    db.session.commit()
    flash("تم تحديث المستخدم بنجاح", "success")
    return redirect(url_for("users"))


@app.route("/account/password", methods=["GET", "POST"])
def change_password():
    user = g.current_user
    if request.method == "POST":
        current_password = request.form.get("current_password") or ""
        new_password = request.form.get("new_password") or ""
        confirm_password = request.form.get("confirm_password") or ""
        if not user.check_password(current_password):
            flash("كلمة المرور الحالية غير صحيحة", "danger")
            return redirect(url_for("change_password"))
        if len(new_password) < 6:
            flash("كلمة المرور الجديدة يجب أن تكون 6 أحرف على الأقل", "danger")
            return redirect(url_for("change_password"))
        if new_password == "123456":
            flash("لا يمكن استخدام كلمة المرور الافتراضية 123456", "danger")
            return redirect(url_for("change_password"))
        if new_password != confirm_password:
            flash("تأكيد كلمة المرور غير متطابق", "danger")
            return redirect(url_for("change_password"))
        if new_password == current_password:
            flash("اختر كلمة مرور مختلفة عن الحالية", "danger")
            return redirect(url_for("change_password"))
        user.set_password(new_password)
        db.session.commit()
        session.pop("must_change_password", None)
        flash("تم تغيير كلمة المرور بنجاح", "success")
        return redirect(url_for("index"))
    return render_template("change_password.html")


@app.route("/")
def index():
    projects = Project.query.count()
    accounts = ChartOfAccount.query.count()
    progress_payments = ProgressPayment.query.count()
    journal_entries = JournalEntry.query.count()
    subcontractors = Subcontractor.query.count()
    suppliers = Supplier.query.count()
    purchase_orders = PurchaseOrder.query.count()
    inventory_transactions = InventoryTransaction.query.count()
    labor_entries = LaborEntry.query.count()
    equipment_items = Equipment.query.count()
    draft_purchase_orders = PurchaseOrder.query.filter(PurchaseOrder.status == "مفتوح").count()
    paid_purchase_orders = PurchaseOrder.query.filter(PurchaseOrder.status == "مدفوع").count()
    outstanding_payments = ProgressPayment.query.filter(ProgressPayment.net_value > 0).count()
    draft_journal_entries = JournalEntry.query.filter(JournalEntry.status != "مرحل").count()
    # قيمة القيود المرحلة فقط، لأن المسودات لا تُعد حركة محاسبية معتمدة
    total_journal_value = round(
        sum(as_float(item.amount) for item in JournalEntry.query.filter(JournalEntry.status == "مرحل").all()),
        2,
    )
    recent_entries = JournalEntry.query.order_by(JournalEntry.id.desc()).limit(8).all()
    all_accounts = ChartOfAccount.query.order_by(ChartOfAccount.category, ChartOfAccount.code).all()
    account_balances = build_account_balances(all_accounts)

    # حساب المورد/المقاول دائن بطبيعته: الرصيد السالب في نظام (مدين - دائن) يعني مستحقًا عليه للشركة.
    # حساب العميل مدين بطبيعته: الرصيد الموجب هو المستحق لنا لدى العميل.
    supplier_dues_rows = []
    customer_dues_rows = []
    for account in all_accounts:
        balance = as_float(account_balances.get(account.id, 0.0))
        if account.category in ("الموردين", "موردين", "مقاولي الباطن"):
            due_amount = -balance if balance < 0 else 0.0
            if due_amount > 0:
                supplier_dues_rows.append({"name": account.name, "amount": round(due_amount, 2)})
        elif account.category == "العملاء":
            due_amount = balance if balance > 0 else 0.0
            if due_amount > 0:
                customer_dues_rows.append({"name": account.name, "amount": round(due_amount, 2)})

    top_supplier_dues = sorted(supplier_dues_rows, key=lambda row: row["amount"], reverse=True)[:5]
    top_customer_dues = sorted(customer_dues_rows, key=lambda row: row["amount"], reverse=True)[:5]

    # الإيرادات الفعلية = صافي الدائن على حسابات الإيرادات (القيود المرحلة فقط)
    revenue_total = 0.0
    for account in all_accounts:
        if account.category in ("الإيرادات", "فروق أسعار"):
            revenue_total += -as_float(account_balances.get(account.id, 0.0))
    revenue_total = round(max(revenue_total, 0.0), 2)

    # إيراد المبيعات = مستخلصات العملاء فقط (بدون مقاولي باطن)
    client_certificates = client_progress_query().all()
    orders_sales_total = round(sum(as_float(item.total_value) for item in client_certificates), 2)
    collections_total = round(sum(as_float(item.amount) for item in ClientReceipt.query.all()), 2)
    direct_sales_total = round(max(revenue_total - orders_sales_total, 0.0), 2)
    sales_total = orders_sales_total + direct_sales_total
    orders_sales_ratio = (orders_sales_total / sales_total) if sales_total > 0 else 0.0
    direct_sales_ratio = (direct_sales_total / sales_total) if sales_total > 0 else 0.0

    active_custody_rows = build_custody_balances(CustodySettlement.query.all())
    active_custody_total = round(sum(row["remaining"] for row in active_custody_rows if row["remaining"] > 0), 2)
    subcontractor_net_due = round(sum(
        build_subcontractor_statement(item)["net_due"] for item in Subcontractor.query.all()
    ), 2)

    journal_boards = [
        {
            "title": "مستخلصات العملاء",
            "subtitle": "إجمالي الأعمال المفوترة للعميل",
            "count": len(client_certificates),
            "total": orders_sales_total,
            "primary_action": {"label": "مستخلص عميل", "url": url_for("client_progress_payments")},
            "secondary_action": {"label": "مستخلص باطن", "url": url_for("progress_payments")},
            "theme": "sales",
        },
        {
            "title": "المقايسات",
            "subtitle": "قيمة مقايسات العملاء النهائية",
            "count": Estimation.query.count(),
            "total": round(sum(as_float(item.final_value) for item in Estimation.query.all()), 2),
            "primary_action": {"label": "مقايسة جديدة", "url": url_for("estimations")},
            "secondary_action": {"label": "ربحية المقايسات", "url": url_for("estimation_profitability_report")},
            "theme": "sales",
        },
        {
            "title": "مشتريات",
            "subtitle": "أوامر شراء وفواتير موردين",
            "count": purchase_orders,
            "total": sum(item.total_value for item in PurchaseOrder.query.all()),
            "primary_action": {"label": "أمر شراء جديد", "url": url_for("purchase_orders")},
            "secondary_action": {"label": "الموردون", "url": url_for("suppliers")},
            "theme": "purchase",
        },
        {
            "title": "بنك وصندوق",
            "subtitle": "حركة الأموال والسيولة",
            "count": journal_entries,
            "total": total_journal_value,
            "primary_action": {"label": "قيد يومية", "url": url_for("journal")},
            "secondary_action": {"label": "دليل الحسابات", "url": url_for("accounts")},
            "theme": "bank",
        },
        {
            "title": "عمليات",
            "subtitle": "المخزون، العمالة، المعدات",
            "count": inventory_transactions + labor_entries + equipment_items,
            "total": 0,
            "primary_action": {"label": "حركة مخزون", "url": url_for("inventory")},
            "secondary_action": {"label": "تقارير المشروع", "url": url_for("project_report")},
            "theme": "ops",
        },
    ]

    return render_template(
        "index.html",
        projects=projects,
        accounts=accounts,
        progress_payments=progress_payments,
        journal_entries=journal_entries,
        draft_journal_entries=draft_journal_entries,
        subcontractors=subcontractors,
        suppliers=suppliers,
        purchase_orders=purchase_orders,
        inventory_transactions=inventory_transactions,
        labor_entries=labor_entries,
        equipment_items=equipment_items,
        draft_purchase_orders=draft_purchase_orders,
        paid_purchase_orders=paid_purchase_orders,
        outstanding_payments=outstanding_payments,
        total_journal_value=total_journal_value,
        recent_entries=recent_entries,
        journal_boards=journal_boards,
        top_supplier_dues=top_supplier_dues,
        top_customer_dues=top_customer_dues,
        orders_sales_total=orders_sales_total,
        collections_total=collections_total,
        direct_sales_total=direct_sales_total,
        orders_sales_ratio=orders_sales_ratio,
        direct_sales_ratio=direct_sales_ratio,
        revenue_total=revenue_total,
        active_custody_total=active_custody_total,
        subcontractor_net_due=subcontractor_net_due,
    )


@app.route("/purchase_orders", methods=["GET", "POST"])
def purchase_orders():
    projects = Project.query.order_by(Project.code).all()
    suppliers = Supplier.query.order_by(Supplier.name).all()
    material_names = get_material_names()
    warehouse_names = get_warehouse_names()
    if request.method == "POST":
        quantity = as_float(request.form.get("quantity"))
        unit_price = as_float(request.form.get("unit_price"))
        discount = as_float(request.form.get("discount"))
        total_value = calculate_purchase_order_total(quantity, unit_price, discount)
        order = PurchaseOrder(
            project_id=as_int(request.form.get("project_id")),
            supplier_id=as_int(request.form.get("supplier_id")),
            item_name=request.form.get("item_name"),
            warehouse_name=request.form.get("warehouse_name"),
            quantity=quantity,
            unit_price=unit_price,
            discount=discount,
            order_number=request.form.get("order_number"),
            invoice_number=request.form.get("invoice_number"),
            date=request.form.get("date") or date.today().isoformat(),
            status=request.form.get("status"),
            total_value=total_value,
            notes=request.form.get("notes"),
        )
        try:
            assert_period_open(order.date)
        except PeriodClosedError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("purchase_orders"))
        db.session.add(order)
        db.session.commit()

        # حركة مخزون تلقائية: أي أمر شراء يضيف الكمية للمخزون.
        sync_purchase_order_to_inventory(order)

        # قيد أمر الشراء يتم مزامنته تلقائيًا (إنشاء/تحديث لنفس القيد).
        try:
            sync_purchase_order_journal(order)
        except PeriodClosedError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("purchase_orders"))

        uploaded = request.files.get("attachment")
        if uploaded and uploaded.filename:
            from routes.cycle import save_attachment
            try:
                save_attachment("purchase_order", order.id, uploaded)
            except ValueError as exc:
                flash(str(exc), "danger")

        flash("تم حفظ أمر الشراء بنجاح", "success")
        return redirect(url_for("purchase_orders"))
    items = PurchaseOrder.query.order_by(PurchaseOrder.date.desc()).all()
    grouped_attachments = {}
    for row in DocumentAttachment.query.filter_by(entity_type="purchase_order").order_by(DocumentAttachment.id.desc()).all():
        grouped_attachments.setdefault(row.entity_id, []).append(row)
    return render_template(
        "purchase_orders.html",
        items=items,
        projects=projects,
        suppliers=suppliers,
        material_names=material_names,
        warehouse_names=warehouse_names,
        today_date=date.today().isoformat(),
        attachments=grouped_attachments,
    )


@app.route("/purchase_orders/<int:order_id>/update", methods=["POST"])
def update_purchase_order(order_id):
    order = PurchaseOrder.query.get_or_404(order_id)
    order.project_id = as_int(request.form.get("project_id")) or order.project_id
    order.supplier_id = as_int(request.form.get("supplier_id"))
    order.item_name = request.form.get("item_name")
    order.warehouse_name = request.form.get("warehouse_name")
    order.quantity = as_float(request.form.get("quantity"))
    order.unit_price = as_float(request.form.get("unit_price"))
    order.discount = as_float(request.form.get("discount"))
    order.order_number = request.form.get("order_number")
    order.invoice_number = request.form.get("invoice_number")
    order.date = request.form.get("date") or None
    order.status = request.form.get("status")
    order.total_value = calculate_purchase_order_total(order.quantity, order.unit_price, order.discount)
    order.notes = request.form.get("notes")
    try:
        assert_period_open(order.date)
    except PeriodClosedError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("purchase_orders"))
    db.session.commit()

    # مزامنة حركة المخزون التلقائية عند تعديل أمر الشراء.
    sync_purchase_order_to_inventory(order)

    # مزامنة قيد أمر الشراء التلقائي عند التعديل.
    try:
        sync_purchase_order_journal(order)
    except PeriodClosedError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("purchase_orders"))

    flash("تم تحديث أمر الشراء بنجاح", "success")
    return redirect(url_for("purchase_orders"))


@app.route("/inventory", methods=["GET", "POST"])
def inventory():
    projects = Project.query.order_by(Project.code).all()
    suppliers = Supplier.query.order_by(Supplier.name).all()
    material_names = get_material_names()
    if request.method == "POST":
        destination_warehouse = request.form.get("destination_warehouse") or ""
        project_id = as_int(request.form.get("project_id"))
        if not project_id:
            flash("يرجى اختيار مشروع صحيح", "danger")
            return redirect(url_for("inventory"))
        transaction = InventoryTransaction(
            project_id=project_id,
            supplier_id=as_int(request.form.get("supplier_id")),
            warehouse_name=request.form.get("warehouse_name"),
            material_name=request.form.get("material_name"),
            quantity=as_float(request.form.get("quantity")),
            unit_cost=as_float(request.form.get("unit_cost")),
            transaction_type=request.form.get("transaction_type"),
            date=request.form.get("date") or date.today().isoformat(),
            notes=request.form.get("notes") or "",
        )
        if request.form.get("transaction_type") == "تحويل" and destination_warehouse:
            transaction.destination_warehouse = destination_warehouse
            transaction.notes = f"تحويل إلى {destination_warehouse} " + transaction.notes
        try:
            assert_period_open(transaction.date)
        except PeriodClosedError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("inventory"))
        db.session.add(transaction)
        db.session.commit()

        try:
            sync_inventory_transaction_journal(transaction)
        except PeriodClosedError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("inventory"))

        flash("تم تسجيل حركة المخزون بنجاح مع قيدها المحاسبي", "success")
        return redirect(url_for("inventory"))
    items = InventoryTransaction.query.order_by(InventoryTransaction.date.desc()).all()
    return render_template(
        "inventory.html",
        items=items,
        projects=projects,
        suppliers=suppliers,
        material_names=material_names,
        today_date=date.today().isoformat(),
    )


@app.route("/inventory/<int:transaction_id>/update", methods=["POST"])
def update_inventory_transaction(transaction_id):
    item = InventoryTransaction.query.get_or_404(transaction_id)
    item.project_id = as_int(request.form.get("project_id")) or item.project_id
    item.supplier_id = as_int(request.form.get("supplier_id"))
    item.warehouse_name = request.form.get("warehouse_name")
    item.destination_warehouse = request.form.get("destination_warehouse") or None
    item.material_name = request.form.get("material_name")
    item.quantity = as_float(request.form.get("quantity"))
    item.unit_cost = as_float(request.form.get("unit_cost"))
    item.transaction_type = request.form.get("transaction_type")
    item.date = request.form.get("date") or None
    item.notes = request.form.get("notes")
    try:
        assert_period_open(item.date)
        db.session.commit()
        sync_inventory_transaction_journal(item)
    except PeriodClosedError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
        return redirect(url_for("inventory"))
    flash("تم تحديث حركة المخزون وقيدها المحاسبي بنجاح", "success")
    return redirect(url_for("inventory"))


@app.route("/custody_settlements", methods=["GET", "POST"])
def custody_settlements():
    sync_journal_related_accounts()
    projects = Project.query.order_by(Project.code).all()
    accounts = ChartOfAccount.query.order_by(ChartOfAccount.category, ChartOfAccount.code).all()
    treasury_accounts = [account for account in accounts if is_treasury_account(account)]
    default_treasury_account_id = treasury_accounts[0].id if treasury_accounts else None
    entity_accounts = get_custody_entity_accounts(accounts)
    expense_accounts = get_custody_expense_accounts(accounts, entity_accounts)
    operation_filter = (request.args.get("operation_type") or "").strip()
    from_date = (request.args.get("from_date") or "").strip()
    to_date = (request.args.get("to_date") or "").strip()
    search_name = (request.args.get("search_name") or request.args.get("entity_name") or "").strip()

    if request.method == "POST":
        operation_type = (request.form.get("operation_type") or request.form.get("voucher_type") or "صرف عهدة").strip()
        if operation_type not in CUSTODY_OPERATION_TYPES:
            operation_type = "صرف عهدة"
        settlement_lines = parse_custody_settlement_lines(request.form)
        amount = as_float(request.form.get("amount"))
        if operation_type == "تسوية عهدة":
            amount = round(sum(line["amount"] for line in settlement_lines), 2)
        if amount <= 0:
            flash("يرجى إدخال مبلغ أكبر من صفر", "danger")
            return redirect(url_for("custody_settlements"))

        entity_name = (request.form.get("entity_name") or "").strip()
        entity_type = (request.form.get("entity_type") or "").strip()
        entity_account_id = as_int(request.form.get("entity_account_id"))
        entity_account = ChartOfAccount.query.get(entity_account_id) if entity_account_id else None

        # إنشاء صاحب عهدة جديد (سائق/مندوب/معدة) مباشرة من الشاشة
        new_owner_name = (request.form.get("new_owner_name") or "").strip()
        new_owner_type = (request.form.get("new_owner_type") or "").strip()
        if not entity_account and new_owner_name:
            entity_account = get_or_create_custody_owner_account(new_owner_type or "سائق", new_owner_name)
            if entity_account:
                db.session.commit()
                entity_account_id = entity_account.id
                entity_type = new_owner_type or entity_type
                entity_name = entity_name or entity_account.name

        if not entity_account and entity_name:
            entity_account = ChartOfAccount.query.filter_by(name=entity_name).first()
            if entity_account:
                entity_account_id = entity_account.id
        if not entity_account_id:
            flash("يرجى اختيار اسم صاحب عهدة من الحسابات المتاحة أو إضافة صاحب عهدة جديد", "danger")
            return redirect(url_for("custody_settlements", operation_type=operation_type))

        if not entity_type and entity_account:
            entity_type = get_custody_owner_type(entity_account) or "سائق"
        if entity_type not in CUSTODY_OWNER_TYPES:
            entity_type = "سائق"
        if not entity_name and entity_account:
            entity_name = (entity_account.name or "").strip()

        expense_nature = (request.form.get("expense_nature") or "").strip()
        settlement = CustodySettlement(
            date=request.form.get("date") or date.today().isoformat(),
            project_id=as_int(request.form.get("project_id")),
            entity_type=entity_type,
            entity_name=entity_name,
            expense_item=(request.form.get("expense_item") or "").strip() or None,
            expense_nature=expense_nature if expense_nature in CUSTODY_EXPENSE_NATURES else None,
            voucher_type="رد" if operation_type == "رد باقي عهدة" else "صرف",
            operation_type=operation_type,
            reference=(request.form.get("reference") or "").strip() or None,
            treasury_account_id=as_int(request.form.get("treasury_account_id")) or default_treasury_account_id,
            entity_account_id=entity_account_id,
            amount=amount,
            settlement_lines=json.dumps(settlement_lines, ensure_ascii=False),
            notes=request.form.get("notes"),
        )
        try:
            assert_period_open(settlement.date)
        except PeriodClosedError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("custody_settlements"))
        db.session.add(settlement)
        db.session.commit()
        try:
            sync_custody_settlement_journal(settlement)
        except PeriodClosedError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("custody_settlements"))

        owner_settlements = CustodySettlement.query.filter_by(entity_account_id=entity_account_id).all()
        owner_balances = build_custody_balances(owner_settlements)
        remaining = owner_balances[0]["remaining"] if owner_balances else 0.0
        flash(
            f"تم حفظ {operation_type} بمبلغ {format_grouped_number(amount)} - "
            f"المتبقي من العهدة: {format_grouped_number(remaining)}",
            "success",
        )
        return redirect(url_for(
            "custody_settlements",
            search_name=(settlement.entity_name or "").strip(),
            operation_type=request.args.get("operation_type") or "",
        ))

    all_items = CustodySettlement.query.order_by(CustodySettlement.date.asc(), CustodySettlement.id.asc()).all()
    all_accounts = ChartOfAccount.query.order_by(ChartOfAccount.category, ChartOfAccount.code).all()

    active_entity_account = None
    if search_name:
        search_lower = search_name.lower()
        active_entity_account = next((account for account in all_accounts if search_lower in (account.name or "").lower()), None)

    account_balances = build_account_balances(all_accounts)
    all_owner_rows = build_custody_owner_rows(all_accounts, all_items, account_balances)
    selected_owner = None
    if active_entity_account:
        selected_owner = next((row for row in all_owner_rows if row["account"].id == active_entity_account.id), None)

    selected_entity_name = selected_owner["entity_name"] if selected_owner else search_name
    selected_entity_type = selected_owner["entity_type"] if selected_owner else "سائق"

    selected_rows = []
    if selected_owner:
        running_balance = 0.0
        for item in selected_owner["settlements"]:
            settlement = item["settlement"]
            op_type = item["operation_type"]
            if operation_filter and operation_filter != op_type:
                continue
            if from_date and (settlement.date or "") < from_date:
                continue
            if to_date and (settlement.date or "") > to_date:
                continue

            amount = as_float(item["amount"])
            debit = amount if op_type in CUSTODY_DISBURSE_OPERATIONS else 0.0
            credit = amount if op_type in ("تسوية عهدة", "رد باقي عهدة") else 0.0
            running_balance += debit - credit

            selected_rows.append({
                "item": settlement,
                "operation_type": op_type,
                "amount": amount,
                "debit": debit,
                "credit": credit,
                "running_balance": round(running_balance, 2),
            })

    selected_totals = {
        "spent": round(sum(row["debit"] for row in selected_rows), 2),
        "settled": round(sum(row["amount"] for row in selected_rows if row["operation_type"] == "تسوية عهدة"), 2),
        "returned": round(sum(row["amount"] for row in selected_rows if row["operation_type"] == "رد باقي عهدة"), 2),
        "balance": selected_owner["balance"] if selected_owner else 0.0,
        "remaining": selected_owner["remaining"] if selected_owner else 0.0,
        "count": len(selected_rows),
    }

    owner_summaries = all_owner_rows
    all_entity_names = [row["entity_name"] for row in all_owner_rows]

    return render_template(
        "custody_settlements.html",
        items=all_items,
        owner_summaries=owner_summaries,
        selected_rows=selected_rows,
        selected_entity_type=selected_entity_type,
        selected_entity_name=selected_entity_name,
        selected_totals=selected_totals,
        all_entity_names=all_entity_names,
        projects=projects,
        treasury_accounts=treasury_accounts,
        entity_accounts=entity_accounts,
        expense_accounts=expense_accounts,
        operation_filter=operation_filter,
        from_date=from_date,
        to_date=to_date,
        accounts=accounts,
        owner_types=CUSTODY_OWNER_TYPES,
        expense_natures=CUSTODY_EXPENSE_NATURES,
        operation_types=CUSTODY_OPERATION_TYPES,
        today_date=date.today().isoformat(),
    )


@app.route("/custody_settlements/export")
def export_custody_settlements():
    sync_journal_related_accounts()
    accounts = ChartOfAccount.query.order_by(ChartOfAccount.category, ChartOfAccount.code).all()
    search_name = (request.args.get("search_name") or request.args.get("entity_name") or "").strip()
    operation_filter = (request.args.get("operation_type") or "").strip()
    from_date = (request.args.get("from_date") or "").strip()
    to_date = (request.args.get("to_date") or "").strip()

    all_items = CustodySettlement.query.order_by(CustodySettlement.date.asc(), CustodySettlement.id.asc()).all()
    account_balances = build_account_balances(accounts)
    owner_rows = build_custody_owner_rows(accounts, all_items, account_balances)
    search_lower = search_name.lower()
    active_entity_account = next((account for account in accounts if search_lower in (account.name or "").lower()), None)
    selected_owner = None
    if active_entity_account:
        selected_owner = next((row for row in owner_rows if row["account"].id == active_entity_account.id), None)

    selected_rows = []
    if selected_owner:
        running_balance = 0.0
        for item in selected_owner["settlements"]:
            settlement = item["settlement"]
            op_type = item["operation_type"]
            if operation_filter and operation_filter != op_type:
                continue
            if from_date and (settlement.date or "") < from_date:
                continue
            if to_date and (settlement.date or "") > to_date:
                continue

            amount = as_float(item["amount"])
            debit = amount if op_type in CUSTODY_DISBURSE_OPERATIONS else 0.0
            credit = amount if op_type in ("تسوية عهدة", "رد باقي عهدة") else 0.0
            running_balance += debit - credit

            selected_rows.append({
                "settlement": settlement,
                "operation_type": op_type,
                "debit": debit,
                "credit": credit,
                "running_balance": running_balance,
            })

    final_balance = selected_owner["balance"] if selected_owner else 0.0

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "كشف العهد"

    # Header info
    sheet.append(["كشف حساب العهدة"])
    sheet.append([f"صاحب العهدة: {search_name or 'كل الحسابات'}"])
    sheet.append([f"نوع العملية: {operation_filter or 'كل العمليات'}"])
    sheet.append([f"الفترة: {from_date or '-'} إلى {to_date or '-'}"])
    sheet.append([])

    header_row = 6
    headers = ["رقم العملية", "نوع العملية", "المرجع", "التاريخ", "اسم صاحب العهدة", "المدين", "الدائن", "الرصيد", "ملاحظات"]
    sheet.append(headers)

    for row in selected_rows:
        settlement = row["settlement"]
        sheet.append([
            f"CUS-{settlement.id:06d}",
            row["operation_type"],
            settlement.reference or "-",
            settlement.date or "-",
            settlement.entity_name or "-",
            row["debit"],
            row["credit"],
            row["running_balance"],
            settlement.notes or "-",
        ])

    total_row_index = sheet.max_row + 1
    sheet.append(["", "", "", "", "إجمالي الرصيد الفعلي للحساب بعد كل الحركات المالية", "", "", final_balance, ""])

    # Styling
    title_cell = sheet.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="right")

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx in range(header_row + 1, total_row_index):
        sheet.cell(row=row_idx, column=6).number_format = "#,##0.00"
        sheet.cell(row=row_idx, column=7).number_format = "#,##0.00"
        sheet.cell(row=row_idx, column=8).number_format = "#,##0.00"

    for col, width in {
        1: 16,
        2: 16,
        3: 18,
        4: 14,
        5: 28,
        6: 14,
        7: 14,
        8: 14,
        9: 30,
    }.items():
        sheet.column_dimensions[chr(64 + col)].width = width

    sheet.freeze_panes = "A7"

    if selected_rows:
        table_end_row = total_row_index - 1
        table_ref = f"A{header_row}:I{table_end_row}"
        table = Table(displayName="CustodyStatementTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    total_fill = PatternFill("solid", fgColor="FFF2CC")
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=total_row_index, column=col)
        cell.fill = total_fill
        if col in (5, 8):
            cell.font = Font(bold=True)
    sheet.cell(row=total_row_index, column=8).number_format = "#,##0.00"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"custody_statement_{(search_name or 'all').replace(' ', '_')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/custody_settlements/<int:settlement_id>/update", methods=["POST"])
def update_custody_settlement(settlement_id):
    settlement = CustodySettlement.query.get_or_404(settlement_id)
    operation_type = (request.form.get("operation_type") or request.form.get("voucher_type") or settlement.operation_type or "صرف عهدة").strip()
    settlement_lines = parse_custody_settlement_lines(request.form)
    amount = as_float(request.form.get("amount"))
    if operation_type == "تسوية عهدة":
        amount = sum(line["amount"] for line in settlement_lines)
    if amount <= 0:
        flash("يرجى إدخال مبلغ أكبر من صفر", "danger")
        return redirect(url_for("custody_settlements"))

    settlement.date = request.form.get("date") or settlement.date
    settlement.project_id = as_int(request.form.get("project_id"))
    settlement.entity_type = request.form.get("entity_type") or settlement.entity_type
    settlement.entity_name = (request.form.get("entity_name") or "").strip()
    settlement.expense_item = (request.form.get("expense_item") or "").strip() or None
    expense_nature = (request.form.get("expense_nature") or "").strip()
    settlement.expense_nature = expense_nature if expense_nature in CUSTODY_EXPENSE_NATURES else None
    settlement.voucher_type = "رد" if operation_type == "رد باقي عهدة" else "صرف"
    settlement.operation_type = operation_type
    settlement.reference = (request.form.get("reference") or "").strip() or None
    settlement.treasury_account_id = as_int(request.form.get("treasury_account_id")) or settlement.treasury_account_id
    settlement.entity_account_id = as_int(request.form.get("entity_account_id")) or settlement.entity_account_id
    settlement.amount = amount
    settlement.settlement_lines = json.dumps(settlement_lines, ensure_ascii=False)
    settlement.notes = request.form.get("notes")
    try:
        assert_period_open(settlement.date)
        db.session.commit()
        sync_custody_settlement_journal(settlement)
    except PeriodClosedError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
        return redirect(url_for("custody_settlements"))
    flash("تم تحديث مستند العهد بنجاح", "success")
    return redirect(url_for("custody_settlements"))


@app.route("/labor", methods=["GET", "POST"])
def labor():
    projects = Project.query.order_by(Project.code).all()
    if request.method == "POST":
        labor = LaborEntry(
            project_id=as_int(request.form.get("project_id")),
            date=request.form.get("date") or date.today().isoformat(),
            description=request.form.get("description"),
            hours=as_float(request.form.get("hours")),
            amount=as_float(request.form.get("amount")),
            advances=as_float(request.form.get("advances")),
            deductions=as_float(request.form.get("deductions")),
        )
        try:
            assert_period_open(labor.date)
        except PeriodClosedError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("labor"))
        db.session.add(labor)
        db.session.commit()
        try:
            sync_labor_journal(labor)
        except PeriodClosedError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("labor"))
        flash("تم تسجيل العمالة وإنشاء القيد التلقائي", "success")
        return redirect(url_for("labor"))
    entries = LaborEntry.query.order_by(LaborEntry.date.desc()).all()
    return render_template("labor.html", entries=entries, projects=projects, today_date=date.today().isoformat())


@app.route("/driver_compensation", methods=["GET", "POST"])
def driver_compensation():
    sync_journal_related_accounts()
    projects = Project.query.order_by(Project.code).all()
    accounts = ChartOfAccount.query.order_by(ChartOfAccount.category, ChartOfAccount.code).all()
    treasury_accounts = [account for account in accounts if is_treasury_account(account)]

    if request.method == "POST":
        driver_name = (request.form.get("driver_name") or "").strip()
        settlement_basis = request.form.get("settlement_basis") or "يومية"
        units = as_float(request.form.get("units"))
        unit_rate = as_float(request.form.get("unit_rate"))
        paid_amount = as_float(request.form.get("paid_amount"))
        gross_amount = max(units * unit_rate, 0.0)
        treasury_account_id = as_int(request.form.get("treasury_account_id"))

        if not driver_name:
            flash("يرجى إدخال اسم السائق", "danger")
            return redirect(url_for("driver_compensation"))
        if settlement_basis not in ("يومية", "نقلة"):
            settlement_basis = "يومية"
        if units < 0 or unit_rate < 0:
            flash("لا يمكن إدخال قيم سالبة للأيام/النقلات أو السعر", "danger")
            return redirect(url_for("driver_compensation"))
        if paid_amount < 0:
            flash("لا يمكن إدخال قيمة سداد سالبة", "danger")
            return redirect(url_for("driver_compensation"))
        if gross_amount <= 0 and paid_amount <= 0:
            flash("يرجى إدخال استحقاق أو سداد بقيمة أكبر من صفر", "danger")
            return redirect(url_for("driver_compensation"))
        if paid_amount > 0 and not treasury_account_id:
            flash("يرجى اختيار حساب الخزنة عند إدخال مبلغ مسدد", "danger")
            return redirect(url_for("driver_compensation"))

        entry = DriverCompensationEntry(
            date=request.form.get("date") or date.today().isoformat(),
            project_id=as_int(request.form.get("project_id")),
            driver_name=driver_name,
            settlement_basis=settlement_basis,
            units=units,
            unit_rate=unit_rate,
            gross_amount=gross_amount,
            paid_amount=paid_amount,
            treasury_account_id=treasury_account_id,
            reference=(request.form.get("reference") or "").strip() or None,
            notes=request.form.get("notes"),
        )
        db.session.add(entry)
        db.session.commit()
        try:
            assert_period_open(entry.date)
            sync_driver_compensation_journals(entry)
        except PeriodClosedError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("driver_compensation"))
        flash("تم حفظ حركة السائق بنجاح", "success")
        return redirect(url_for("driver_compensation"))

    from_date = (request.args.get("from_date") or "").strip()
    to_date = (request.args.get("to_date") or "").strip()
    driver_filter = (request.args.get("driver_name") or "").strip()

    query = DriverCompensationEntry.query
    if from_date:
        query = query.filter(DriverCompensationEntry.date >= from_date)
    if to_date:
        query = query.filter(DriverCompensationEntry.date <= to_date)
    if driver_filter:
        query = query.filter(DriverCompensationEntry.driver_name == driver_filter)
    entries = query.order_by(DriverCompensationEntry.date.desc(), DriverCompensationEntry.id.desc()).all()

    all_driver_names = sorted({
        (value[0] or "").strip()
        for value in DriverCompensationEntry.query.with_entities(DriverCompensationEntry.driver_name).all()
        if (value[0] or "").strip()
    })

    summary_map = defaultdict(lambda: {
        "driver_name": "",
        "daily_units": 0.0,
        "trip_units": 0.0,
        "gross_total": 0.0,
        "paid_total": 0.0,
        "due_total": 0.0,
    })
    for item in entries:
        row = summary_map[item.driver_name]
        row["driver_name"] = item.driver_name
        if item.settlement_basis == "يومية":
            row["daily_units"] += as_float(item.units)
        else:
            row["trip_units"] += as_float(item.units)
        row["gross_total"] += as_float(item.gross_amount)
        row["paid_total"] += as_float(item.paid_amount)
        row["due_total"] = row["gross_total"] - row["paid_total"]

    # ربط الإنتاجية بمصروفات العهد اليومية ومصاريف النقلة (SRS 3.2)
    custody_rows = build_custody_balances(
        CustodySettlement.query.filter(CustodySettlement.entity_type == "سائق").all()
    )
    custody_by_driver = {}
    for custody_row in custody_rows:
        clean_name = (custody_row["entity_name"] or "").replace("سائق - ", "").strip()
        custody_by_driver[clean_name] = custody_row

    for row in summary_map.values():
        custody_row = custody_by_driver.get((row["driver_name"] or "").strip(), {})
        row["custody_disbursed"] = custody_row.get("disbursed", 0.0)
        row["custody_settled"] = custody_row.get("settled_total", 0.0)
        row["custody_trip_expenses"] = custody_row.get("settled_trip", 0.0)
        row["custody_daily_expenses"] = custody_row.get("settled_daily", 0.0)
        row["custody_returned"] = custody_row.get("returned", 0.0)
        # المتبقي من عهدة السائق = إجمالي العهد المنصرفة - (مصروفات النقلات المعتمدة + المصاريف اليومية)
        row["custody_remaining"] = custody_row.get("remaining", 0.0)
        net_productivity = row["gross_total"] - row["custody_settled"]
        row["net_productivity"] = round(net_productivity, 2)

    summary_rows = sorted(summary_map.values(), key=lambda x: x["driver_name"])
    totals = {
        "gross": round(sum(row["gross_total"] for row in summary_rows), 2),
        "paid": round(sum(row["paid_total"] for row in summary_rows), 2),
        "due": round(sum(row["due_total"] for row in summary_rows), 2),
        "custody_disbursed": round(sum(row["custody_disbursed"] for row in summary_rows), 2),
        "custody_settled": round(sum(row["custody_settled"] for row in summary_rows), 2),
        "custody_remaining": round(sum(row["custody_remaining"] for row in summary_rows), 2),
        "net_productivity": round(sum(row["net_productivity"] for row in summary_rows), 2),
    }

    return render_template(
        "driver_compensation.html",
        entries=entries,
        summary_rows=summary_rows,
        totals=totals,
        projects=projects,
        treasury_accounts=treasury_accounts,
        all_driver_names=all_driver_names,
        from_date=from_date,
        to_date=to_date,
        driver_filter=driver_filter,
        today_date=date.today().isoformat(),
    )


@app.route("/driver_compensation/weekly", methods=["POST"])
def driver_compensation_weekly():
    sync_journal_related_accounts()
    driver_name = (request.form.get("driver_name") or "").strip()
    week_days = as_float(request.form.get("week_days"))
    daily_rate = as_float(request.form.get("daily_rate"))
    paid_amount = as_float(request.form.get("paid_amount"))
    treasury_account_id = as_int(request.form.get("treasury_account_id"))
    gross_amount = max(week_days * daily_rate, 0.0)

    if not driver_name:
        flash("يرجى إدخال اسم السائق للتسوية الأسبوعية", "danger")
        return redirect(url_for("driver_compensation"))
    if week_days <= 0 or daily_rate < 0:
        flash("يرجى إدخال عدد أيام وسعر يومية صحيحين", "danger")
        return redirect(url_for("driver_compensation"))
    if paid_amount < 0:
        flash("لا يمكن إدخال مبلغ سداد سالب", "danger")
        return redirect(url_for("driver_compensation"))
    if paid_amount > 0 and not treasury_account_id:
        flash("يرجى اختيار حساب الخزنة عند إدخال مبلغ مسدد", "danger")
        return redirect(url_for("driver_compensation"))

    entry = DriverCompensationEntry(
        date=request.form.get("date") or date.today().isoformat(),
        project_id=as_int(request.form.get("project_id")),
        driver_name=driver_name,
        settlement_basis="يومية",
        units=week_days,
        unit_rate=daily_rate,
        gross_amount=gross_amount,
        paid_amount=paid_amount,
        treasury_account_id=treasury_account_id,
        reference=(request.form.get("reference") or "").strip() or None,
        notes=(request.form.get("notes") or "").strip() or "تسوية أسبوعية",
    )
    db.session.add(entry)
    db.session.commit()
    try:
        assert_period_open(entry.date)
        sync_driver_compensation_journals(entry)
    except PeriodClosedError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("driver_compensation"))
    flash("تم حفظ التسوية الأسبوعية للسائق بنجاح", "success")
    return redirect(url_for("driver_compensation"))


@app.route("/driver_compensation/<int:entry_id>/update", methods=["POST"])
def update_driver_compensation(entry_id):
    entry = DriverCompensationEntry.query.get_or_404(entry_id)
    driver_name = (request.form.get("driver_name") or "").strip()
    settlement_basis = request.form.get("settlement_basis") or entry.settlement_basis
    units = as_float(request.form.get("units"))
    unit_rate = as_float(request.form.get("unit_rate"))
    paid_amount = as_float(request.form.get("paid_amount"))
    gross_amount = max(units * unit_rate, 0.0)
    treasury_account_id = as_int(request.form.get("treasury_account_id"))

    if not driver_name:
        flash("يرجى إدخال اسم السائق", "danger")
        return redirect(url_for("driver_compensation"))
    if settlement_basis not in ("يومية", "نقلة"):
        settlement_basis = "يومية"
    if units < 0 or unit_rate < 0 or paid_amount < 0:
        flash("الرجاء إدخال قيم صحيحة أكبر أو تساوي صفر", "danger")
        return redirect(url_for("driver_compensation"))
    if paid_amount > 0 and not treasury_account_id:
        flash("يرجى اختيار حساب الخزنة عند إدخال مبلغ مسدد", "danger")
        return redirect(url_for("driver_compensation"))

    entry.date = request.form.get("date") or entry.date
    entry.project_id = as_int(request.form.get("project_id"))
    entry.driver_name = driver_name
    entry.settlement_basis = settlement_basis
    entry.units = units
    entry.unit_rate = unit_rate
    entry.gross_amount = gross_amount
    entry.paid_amount = paid_amount
    entry.treasury_account_id = treasury_account_id
    entry.reference = (request.form.get("reference") or "").strip() or None
    entry.notes = request.form.get("notes")
    db.session.commit()
    try:
        assert_period_open(entry.date)
        sync_driver_compensation_journals(entry)
    except PeriodClosedError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("driver_compensation"))
    flash("تم تحديث حركة السائق بنجاح", "success")
    return redirect(url_for("driver_compensation"))


@app.route("/labor/<int:labor_id>/update", methods=["POST"])
def update_labor_entry(labor_id):
    entry = LaborEntry.query.get_or_404(labor_id)
    entry.project_id = as_int(request.form.get("project_id")) or entry.project_id
    entry.date = request.form.get("date") or None
    entry.description = request.form.get("description")
    entry.hours = as_float(request.form.get("hours"))
    entry.amount = as_float(request.form.get("amount"))
    entry.advances = as_float(request.form.get("advances"))
    entry.deductions = as_float(request.form.get("deductions"))
    db.session.commit()
    try:
        sync_labor_journal(entry)
    except PeriodClosedError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("labor"))
    flash("تم تحديث سجل العمالة والقيد التلقائي", "success")
    return redirect(url_for("labor"))


@app.route("/equipment", methods=["GET", "POST"])
def equipment():
    projects = Project.query.order_by(Project.code).all()
    if request.method == "POST":
        equip = Equipment(
            name=request.form.get("name"),
            purchase_cost=as_float(request.form.get("purchase_cost")),
            operating_cost=as_float(request.form.get("operating_cost")),
            maintenance=as_float(request.form.get("maintenance")),
            hours_used=as_float(request.form.get("hours_used")),
            project_id=as_int(request.form.get("project_id")),
        )
        db.session.add(equip)
        db.session.commit()
        sync_journal_related_accounts()
        try:
            sync_equipment_journals(equip)
        except PeriodClosedError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("equipment"))
        flash("تم حفظ المعدة وإنشاء قيود الشراء والتشغيل", "success")
        return redirect(url_for("equipment"))
    entries = Equipment.query.order_by(Equipment.name).all()
    return render_template("equipment.html", entries=entries, projects=projects)


@app.route("/equipment/<int:equipment_id>/update", methods=["POST"])
def update_equipment(equipment_id):
    item = Equipment.query.get_or_404(equipment_id)
    item.name = request.form.get("name")
    item.purchase_cost = as_float(request.form.get("purchase_cost"))
    item.operating_cost = as_float(request.form.get("operating_cost"))
    item.maintenance = as_float(request.form.get("maintenance"))
    item.hours_used = as_float(request.form.get("hours_used"))
    item.project_id = as_int(request.form.get("project_id"))
    db.session.commit()
    sync_journal_related_accounts()
    try:
        sync_equipment_journals(item)
    except PeriodClosedError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("equipment"))
    flash("تم تحديث المعدة وقيودها التلقائية", "success")
    return redirect(url_for("equipment"))



# ---------------------------------------------------------------------------
# مقايسات العملاء (SRS 4)
# ---------------------------------------------------------------------------


@app.route("/estimations", methods=["GET", "POST"])
def estimations():
    sync_journal_related_accounts()
    projects = Project.query.order_by(Project.code).all()

    if request.method == "POST":
        client_name = resolve_client_name(request.form)
        if not client_name:
            flash("يرجى اختيار عميل مسجّل أو إدخال اسم عميل جديد", "danger")
            return redirect(url_for("estimations"))

        item_lines = parse_estimation_items(request.form)
        if not item_lines:
            flash("يرجى إدخال بند واحد على الأقل في جدول الأعمال والتوريدات", "danger")
            return redirect(url_for("estimations"))

        project_id = as_int(request.form.get("project_id"))
        project = Project.query.get(project_id) if project_id else None
        admin_mode = request.form.get("admin_mode") or "إضافة"
        estimation = Estimation(
            code=(request.form.get("code") or "").strip() or generate_estimation_code(),
            date=request.form.get("date") or date.today().isoformat(),
            client_name=client_name,
            project_id=project_id,
            project_name=(request.form.get("project_name") or "").strip() or (project.project_name if project else None),
            discount_percentage=as_float(request.form.get("discount_percentage")),
            admin_percentage=as_float(request.form.get("admin_percentage")),
            admin_mode=admin_mode if admin_mode in ("إضافة", "خصم") else "إضافة",
            status="مسودة",
            notes=request.form.get("notes"),
        )
        db.session.add(estimation)
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("رقم المقايسة مستخدم بالفعل، اختر رقمًا مختلفًا", "danger")
            return redirect(url_for("estimations"))

        for line in item_lines:
            db.session.add(EstimationItem(estimation_id=estimation.id, **line))
        db.session.commit()

        recalculate_estimation(estimation)
        db.session.commit()

        flash(
            f"تم حفظ المقايسة {estimation.code}: "
            f"الإجمالي {format_grouped_number(estimation.total_value)} - "
            f"الخصم {format_grouped_number(estimation.discount_value)} - "
            f"الصافي بعد الخصم {format_grouped_number(estimation.net_after_discount)} - "
            f"الإداريات {format_grouped_number(estimation.admin_value)} - "
            f"القيمة النهائية {format_grouped_number(estimation.final_value)}",
            "success",
        )
        return redirect(url_for("estimation_detail", estimation_id=estimation.id))

    items = Estimation.query.order_by(Estimation.id.desc()).all()
    totals = {
        "total_value": round(sum(as_float(item.total_value) for item in items), 2),
        "discount_value": round(sum(as_float(item.discount_value) for item in items), 2),
        "net_after_discount": round(sum(as_float(item.net_after_discount) for item in items), 2),
        "admin_value": round(sum(as_float(item.admin_value) for item in items), 2),
        "final_value": round(sum(as_float(item.final_value) for item in items), 2),
    }
    return render_template(
        "estimations.html",
        items=items,
        projects=projects,
        unit_options=UNIT_OPTIONS,
        next_code=generate_estimation_code(),
        today_date=date.today().isoformat(),
        totals=totals,
        client_names=list_client_names(),
    )


@app.route("/estimations/<int:estimation_id>")
def estimation_detail(estimation_id):
    estimation = Estimation.query.get_or_404(estimation_id)
    items = EstimationItem.query.filter_by(estimation_id=estimation.id).order_by(EstimationItem.id).all()
    projects = Project.query.order_by(Project.code).all()
    linked_journals = JournalEntry.query.filter(
        JournalEntry.description.like(f"%EST-AUTO:{estimation.id}%")
    ).order_by(JournalEntry.id.asc()).all()
    return render_template(
        "estimation_detail.html",
        estimation=estimation,
        items=items,
        projects=projects,
        unit_options=UNIT_OPTIONS,
        linked_journals=linked_journals,
        client_names=list_client_names(),
    )


@app.route("/estimations/<int:estimation_id>/update", methods=["POST"])
def update_estimation(estimation_id):
    estimation = Estimation.query.get_or_404(estimation_id)
    estimation.client_name = resolve_client_name(request.form) or estimation.client_name
    estimation.date = request.form.get("date") or estimation.date
    estimation.project_id = as_int(request.form.get("project_id"))
    estimation.project_name = (request.form.get("project_name") or "").strip() or None
    estimation.discount_percentage = as_float(request.form.get("discount_percentage"))
    estimation.admin_percentage = as_float(request.form.get("admin_percentage"))
    admin_mode = request.form.get("admin_mode") or estimation.admin_mode
    estimation.admin_mode = admin_mode if admin_mode in ("إضافة", "خصم") else estimation.admin_mode
    estimation.notes = request.form.get("notes")

    item_lines = parse_estimation_items(request.form)
    if item_lines:
        for existing_item in EstimationItem.query.filter_by(estimation_id=estimation.id).all():
            db.session.delete(existing_item)
        db.session.flush()
        for line in item_lines:
            db.session.add(EstimationItem(estimation_id=estimation.id, **line))

    db.session.commit()
    recalculate_estimation(estimation)
    db.session.commit()
    sync_estimation_journal(estimation)
    flash(
        f"تم تحديث المقايسة - القيمة النهائية {format_grouped_number(estimation.final_value)}",
        "success",
    )
    return redirect(url_for("estimation_detail", estimation_id=estimation.id))


@app.route("/estimations/<int:estimation_id>/status", methods=["POST"])
def update_estimation_status(estimation_id):
    estimation = Estimation.query.get_or_404(estimation_id)
    new_status = (request.form.get("status") or "").strip()
    if new_status not in ("مسودة", "معتمدة", "ملغاة"):
        flash("حالة المقايسة غير صحيحة", "danger")
        return redirect(url_for("estimation_detail", estimation_id=estimation.id))

    estimation.status = new_status
    db.session.commit()
    recalculate_estimation(estimation)
    db.session.commit()
    sync_estimation_journal(estimation)

    if new_status == "معتمدة" and request.form.get("update_contract_value") == "on" and estimation.project:
        estimation.project.contract_value = as_float(estimation.final_value)
        db.session.commit()
        flash("تم تحديث قيمة عقد المشروع بقيمة المقايسة النهائية", "success")

    flash(f"تم تحويل حالة المقايسة إلى {new_status}", "success")
    return redirect(url_for("estimation_detail", estimation_id=estimation.id))


from routes.catalog import register as register_catalog
from routes.cycle import register as register_operating_cycle
from routes.hr import register as register_hr
from routes.journal import register as register_journal
from routes.reports import register as register_reports

register_catalog(app)
register_operating_cycle(app)
register_hr(app)
register_journal(app)
register_reports(app)
